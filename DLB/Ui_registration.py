# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'c:\Users\gulya\Desktop\DLB\registration.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
import Ui_registration_section
from openpyxl import Workbook, load_workbook
from Ui_mainwindow import id_registered


class Ui_Form(object):
    def __init__(self):
        super().__init__()

        self.window = QtWidgets.QWidget()
        self.setupUi(self.window)
        self.window.show()

    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(282, 272)
        self.id_le = QtWidgets.QLineEdit(Form)
        self.id_le.setGeometry(QtCore.QRect(100, 60, 151, 22))
        self.id_le.setObjectName("id_le")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(30, 60, 31, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(30, 90, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.name_le = QtWidgets.QLineEdit(Form)
        self.name_le.setGeometry(QtCore.QRect(100, 90, 151, 22))
        self.name_le.setObjectName("name_le")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(30, 180, 171, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.sectionsNumber_le = QtWidgets.QLineEdit(Form)
        self.sectionsNumber_le.setGeometry(QtCore.QRect(190, 180, 61, 22))
        self.sectionsNumber_le.setObjectName("sectionsNumber_le")
        self.age_le = QtWidgets.QLineEdit(Form)
        self.age_le.setGeometry(QtCore.QRect(100, 150, 151, 22))
        self.age_le.setObjectName("age_le")
        self.label_4 = QtWidgets.QLabel(Form)
        self.label_4.setGeometry(QtCore.QRect(30, 150, 71, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(Form)
        self.label_5.setGeometry(QtCore.QRect(80, 10, 161, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.go_btn = QtWidgets.QPushButton(Form)
        self.go_btn.setGeometry(QtCore.QRect(160, 220, 93, 28))
        self.go_btn.setObjectName("go_btn")
        self.label_6 = QtWidgets.QLabel(Form)
        self.label_6.setGeometry(QtCore.QRect(30, 120, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.radioButton = QtWidgets.QRadioButton(Form)
        self.radioButton.setGeometry(QtCore.QRect(120, 120, 51, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.radioButton.setFont(font)
        self.radioButton.setObjectName("radioButton")
        self.radioButton_2 = QtWidgets.QRadioButton(Form)
        self.radioButton_2.setGeometry(QtCore.QRect(190, 120, 51, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.radioButton_2.setFont(font)
        self.radioButton_2.setObjectName("radioButton_2")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)
        self.go_btn.clicked.connect(self.register_id)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Регистрация"))
        self.label.setText(_translate("Form", "ID:"))
        self.label_2.setText(_translate("Form", "Имя:"))
        self.label_3.setText(_translate("Form", "Количество секций:"))
        self.label_4.setText(_translate("Form", "Возраст:"))
        self.label_5.setText(_translate("Form", "Регистрация"))
        self.go_btn.setText(_translate("Form", "Далее"))
        self.label_6.setText(_translate("Form", "Пол:"))
        self.radioButton.setText(_translate("Form", "М"))
        self.radioButton_2.setText(_translate("Form", "Ж"))

    def open_sectioner(self, line, name):
        self.window_to_open = Ui_registration_section.Ui_Form(line, name)
        self.window.close()

    def register_id(self):
        id_to_register = self.id_le.text()
        id_registered.append(id_to_register)
        line = str(int(id_to_register) + 1)

        #auditory info changing
        wb2 = Workbook()
        wb2 = load_workbook('auditory.xlsx')
        ws2 = wb2.active
        old_value = str(ws2['C2'].value)
        if id_to_register not in old_value:
            string = (old_value + ', ' + id_to_register).replace('None', '')
            ws2['C2'] = string[2:]
        wb2.save('auditory.xlsx')

        #people info changing
        wb = Workbook()
        wb = load_workbook('people.xlsx')
        ws = wb.active
            #gender
        index = 'B' + line
        if self.radioButton.isChecked():
            ws[index] = 'М'
        else:
            ws[index] = 'Ж'
            #age
        index = 'C' + line
        ws[index] = self.age_le.text()
            #name
        name = self.name_le.text()
        index = 'E' + line
        ws[index] = name
            #sections are in the next window
        wb.save('people.xlsx')

        self.open_sectioner(line, name)
