import time
import datetime
from openpyxl import Workbook, load_workbook
from Ui_mainwindow import id_registered

people = {
    'id': "",
    'gender': "",
    'age': "",
    #unknown only
    'visit_dt': "",
    #known only
    'name': "known",
    'sections': [],
}

section = {
    'name': "",
    'registration_dt': "",
    'visit_dt': "",
    'achievements': ""
}

auditory = {
    'name': "",
    'attendance': "",
    'registered_id': "",
    'visitors_id': ""
}

columns_people = list(people.keys())
columns_people.extend(list(people.keys()))
columns_auditory = list(auditory.keys())

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
num = 0

wb = Workbook()
wb = load_workbook('people.xlsx')
ws = wb.active
wb2 = Workbook()
wb2 = load_workbook('auditory.xlsx')
ws2 = wb2.active

line_people = 2 # count_people = line_people - 2
line_auditory = 2 # count_auditory = line_auditory - 2

vector = {
    'id': "",
    'gender': "",
    'age': ""
}

#id_registered = []
id_list = []

attendance = {}

class DB(object):

    def vector_put(vector_to_put):
        global people, section, auditory, wb, ws, wb2, ws2, columns_people, columns_auditory, alphabet, line_people, line_auditory, num
        #global id_registered
        num = 0

        def put_people(parameter):
            #GLOBAL or WITH CLASS?? -> only self. ??
            global alphabet, num, line_people, ws
            index = alphabet[num] + str(line_people)
            ws[index] = (str(ws[index].value) + parameter).replace('None', '') # DON'T KNOW IF it works.
            num += 1

        def put_people_registered(parameter, line):
            global ws
            #'E' is the 5th letter in alphabet
            index = 'E' + str(line)
            ws[index] = (str(ws[index].value) + parameter).replace('None', '')

        def append_visit(id_visitor, timedate):
            global ws2
            timedate = timedate[:len(timedate)-3] # only day hh:mm is left
            day = timedate[:len(timedate)-6] # only day is left
            #visitors_id 
            visitors_id = id_visitor + '-' + timedate + '__'
            ws2['D2'] = (str(ws2['D2'].value) + visitors_id).replace('None', '')
            #attendance
            #attendance_string = str(ws2['B2'].value).replace('None', '') - ne nuzhno
            if day in attendance:
                attendance[day] += 1
            else:
                attendance[day] = 1
            attendance_string = ""
            for day in attendance:
                attendance_string += str(attendance[day]) + "-" + day + "__"
            #ws2['B2'] = (str(ws2['B2'].value) + attendance_string).replace('None', '') # но это очень странно в масштабах 5-минутной презентации
            ws2['B2'] = attendance_string

        #people db
        if int(vector_to_put['id']) == (line_people - 1):
        # ЕСЛИ ЧЕЛОВЕК НОВЫЙ
            if int(vector_to_put['id']) in id_registered:
                #ЕСЛИ ЧЕЛОВЕК ЗАРЕГАН
                pass
            else:
            #ЕСЛИ ЧЕЛОВЕК НЕ ЗАРЕГАН
                put_people(vector_to_put['id'])
                id_list.append(vector_to_put['id'])
                #index = alphabet[num] + str(line_people)
                #ws[index] = vector_to_put['id']
                
                # WHAT will RPi send instead of ""?
                # ЕСЛИ RPi ПРИСЛАЛА ТОЛЬКО id
                if(vector_to_put['gender'] != ""):
                    put_people(vector_to_put['gender'])
                    #index = alphabet[num] + str(line_people)
                    #ws[index] = vector_to_put['gender']
                    put_people(vector_to_put['age'])
                    #index = alphabet[num] + str(line_people)
                    #ws[index] = vector_to_put['age']
                else:
                    num += 2

                timedate = str(datetime.datetime.now())
                timedate = timedate[:len(timedate)-7]
                parameter = 'DLB-' + timedate + '__'
                put_people(parameter)
                append_visit(str(vector_to_put['id']), timedate)


                """
                for num in range(0, 6):
                    index = alphabet[num] + str(line_people)
                    ws[index] = vector_to_put['id']
                """
                line_people += 1
        elif int(vector_to_put['id']) < (line_people - 1):
        # ЕСЛИ ЧЕЛОВЕК УЖЕ ЕСТЬ В БД
            line = int(vector_to_put['id']) + 1

            timedate = str(datetime.datetime.now())
            timedate = timedate[:len(timedate)-7]
            parameter = 'DLB-' + timedate + '__'
            put_people_registered(parameter, line)
            append_visit(str(vector_to_put['id']), timedate)
        else:
        #НЕВОЗМОЖНО, ЕСЛИ id ПОСТУПАЮТ ПО ПОРЯДКУ
            print("blyats")

        #auditory db
        wb.save('people.xlsx')
        wb2.save('auditory.xlsx')

def imitator():
    while(True):
        #global vector
        vector_imitated = {
            'name': "",
            'attendance': "",
            'registered_id': "",
            'visitors_id': ""
        }
        vector_imitated['id'] = input()
        vector_imitated['gender'] = input()
        if vector_imitated['gender'] in ['m', 'f']:
            vector_imitated['age'] = input()
        else:
            vector_imitated['gender'] = ""
        DB.vector_put(vector_imitated)


if __name__ == '__main__':
    imitator()
    #vector_put()
    #или повторяющиеся callback запросы к RPi?