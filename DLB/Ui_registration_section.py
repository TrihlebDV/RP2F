# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'c:\Users\gulya\Desktop\DLB\registration_section.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook, load_workbook
import Ui_registration_success


class Ui_Form(object):
    def __init__(self, line, name):
        super().__init__()

        self.window = QtWidgets.QWidget()
        self.setupUi(self.window)
        self.window.show()

        self.line = line
        self.name = name

    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(630, 344)
        self.section_le = QtWidgets.QLineEdit(Form)
        self.section_le.setGeometry(QtCore.QRect(20, 50, 591, 22))
        self.section_le.setObjectName("section_le")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(270, 80, 211, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.achievements_le = QtWidgets.QLineEdit(Form)
        self.achievements_le.setGeometry(QtCore.QRect(20, 110, 591, 181))
        self.achievements_le.setObjectName("achievements_le")
        self.pushButton_2 = QtWidgets.QPushButton(Form)
        self.pushButton_2.setGeometry(QtCore.QRect(520, 300, 93, 28))
        self.pushButton_2.setObjectName("pushButton_2")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(250, 20, 211, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)
        self.pushButton_2.clicked.connect(self.open_next_window)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Регистрация"))
        self.label_2.setText(_translate("Form", "Достижения"))
        self.pushButton_2.setText(_translate("Form", "Далее"))
        self.label_3.setText(_translate("Form", "Название секции"))

    def open_next_window(self):
        #people info changing: section (yes, only one)
        wb = Workbook()
        wb = load_workbook('people.xlsx')
        ws = wb.active
        index = 'F' + self.line
        ws[index] = '[' + 'DLB' + '-' + self.achievements_le.text() + ']'
        wb.save('people.xlsx')

        self.window_to_open = Ui_registration_success.Ui_Form(self.name)
        self.window.close()
